"""Importación masiva de biblioteca de kits de costos.

Formatos soportados: .xlsx, .csv, .json. La estructura esperada se documenta en
la plantilla que devuelve `build_template_workbook()`.

Reglas de upsert:
- MasterFormat: por `division_code`.
- ActivityKit (maestro, proyecto=NULL): por `codigo_kit`; fallback a `nombre` si
  el código viene vacío.
- Activity (dentro de kit maestro): por `(activity_kit, codigo_actividad)`.
"""
from __future__ import annotations

import csv
import json
import zipfile
from dataclasses import dataclass, field
from decimal import Decimal, InvalidOperation
from io import BytesIO, StringIO
from typing import Any

import tablib
from django.core.files.uploadedfile import UploadedFile
from django.db import transaction
from openpyxl import Workbook, load_workbook

from .models import Activity, ActivityKit, MasterFormat


SHEET_DIVISIONS = "Divisiones"
SHEET_KITS = "Kits"
SHEET_ACTIVITIES = "Actividades"

DIVISION_HEADERS = ["division_code", "division_name"]
KIT_HEADERS = ["codigo_kit", "nombre", "descripcion", "color"]
ACTIVITY_HEADERS = [
    "codigo_kit",
    "codigo_actividad",
    "descripcion",
    "unidad",
    "cu_total",
    "material",
    "mano_obra",
    "equipo",
    "division_code",
]


class ImportError_(Exception):
    """Error de importación que se traduce a HTTP 400."""


@dataclass
class ImportPayload:
    divisions: list[dict[str, Any]] = field(default_factory=list)
    kits: list[dict[str, Any]] = field(default_factory=list)
    activities: list[dict[str, Any]] = field(default_factory=list)

    def extend(self, other: "ImportPayload") -> None:
        self.divisions.extend(other.divisions)
        self.kits.extend(other.kits)
        self.activities.extend(other.activities)


@dataclass
class ImportSummary:
    divisions_created: int = 0
    divisions_updated: int = 0
    kits_created: int = 0
    kits_updated: int = 0
    activities_created: int = 0
    activities_updated: int = 0
    errors: list[dict[str, Any]] = field(default_factory=list)

    def as_dict(self) -> dict[str, Any]:
        return {
            "divisions_created": self.divisions_created,
            "divisions_updated": self.divisions_updated,
            "kits_created": self.kits_created,
            "kits_updated": self.kits_updated,
            "activities_created": self.activities_created,
            "activities_updated": self.activities_updated,
            "errors": self.errors,
        }


# ── Parsing ──────────────────────────────────────────────────────────────────


def parse_file(uploaded: UploadedFile) -> ImportPayload:
    name = (uploaded.name or "").lower()
    content = uploaded.read()
    if name.endswith(".xlsx"):
        return _parse_xlsx(content)
    if name.endswith(".json"):
        return _parse_json(content)
    if name.endswith(".csv"):
        return _parse_csv(content, source_name=uploaded.name)
    raise ImportError_(
        f"Formato no soportado para «{uploaded.name}». Use .xlsx, .csv o .json."
    )


def _parse_xlsx(content: bytes) -> ImportPayload:
    try:
        wb = load_workbook(BytesIO(content), data_only=True, read_only=True)
    except Exception as exc:
        raise ImportError_(f"No se pudo leer el archivo Excel: {exc}") from exc

    payload = ImportPayload()
    if SHEET_DIVISIONS in wb.sheetnames:
        payload.divisions = _rows_from_sheet(wb[SHEET_DIVISIONS], DIVISION_HEADERS)
    if SHEET_KITS in wb.sheetnames:
        payload.kits = _rows_from_sheet(wb[SHEET_KITS], KIT_HEADERS)
    if SHEET_ACTIVITIES in wb.sheetnames:
        payload.activities = _rows_from_sheet(wb[SHEET_ACTIVITIES], ACTIVITY_HEADERS)
    if not (payload.divisions or payload.kits or payload.activities):
        raise ImportError_(
            f"El Excel debe contener al menos una hoja: {SHEET_DIVISIONS!r}, "
            f"{SHEET_KITS!r} o {SHEET_ACTIVITIES!r}."
        )
    return payload


def _rows_from_sheet(sheet, expected_headers: list[str]) -> list[dict[str, Any]]:
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []
    header_row = [str(c).strip().lower() if c is not None else "" for c in rows[0]]
    result = []
    for i, raw_row in enumerate(rows[1:], start=2):
        if all(v is None or str(v).strip() == "" for v in raw_row):
            continue
        row_dict = {}
        for header, value in zip(header_row, raw_row):
            if header in expected_headers:
                row_dict[header] = value
        row_dict["__row"] = i
        result.append(row_dict)
    return result


def _parse_csv(content: bytes, source_name: str = "") -> ImportPayload:
    try:
        text = content.decode("utf-8-sig")
    except UnicodeDecodeError:
        text = content.decode("latin-1")

    dataset = tablib.Dataset().load(text, format="csv")
    headers = [h.strip().lower() for h in dataset.headers or []]
    payload = ImportPayload()
    rows = [dict(zip(headers, row)) for row in dataset]
    for idx, row in enumerate(rows, start=2):
        row["__row"] = idx

    kind = _detect_csv_kind(headers)
    if kind == "divisions":
        payload.divisions = rows
    elif kind == "kits":
        payload.kits = rows
    elif kind == "activities":
        payload.activities = rows
    else:
        raise ImportError_(
            f"No se pudo detectar el tipo del CSV «{source_name}». "
            f"Se esperan encabezados de Divisiones, Kits o Actividades."
        )
    return payload


def _detect_csv_kind(headers: list[str]) -> str | None:
    hset = set(headers)
    if "codigo_actividad" in hset and "cu_total" in hset:
        return "activities"
    if "codigo_kit" in hset and "nombre" in hset and "codigo_actividad" not in hset:
        return "kits"
    if "division_code" in hset and "division_name" in hset and "codigo_actividad" not in hset:
        return "divisions"
    return None


def _parse_json(content: bytes) -> ImportPayload:
    try:
        data = json.loads(content.decode("utf-8-sig"))
    except (UnicodeDecodeError, json.JSONDecodeError) as exc:
        raise ImportError_(f"JSON inválido: {exc}") from exc
    if not isinstance(data, dict):
        raise ImportError_("El JSON raíz debe ser un objeto.")

    payload = ImportPayload()
    for i, d in enumerate(data.get("divisiones", []) or [], start=1):
        payload.divisions.append({**d, "__row": i})

    for i, k in enumerate(data.get("kits", []) or [], start=1):
        actividades = k.pop("actividades", []) or []
        payload.kits.append({**k, "__row": i})
        codigo_kit = k.get("codigo_kit")
        for j, a in enumerate(actividades, start=1):
            payload.activities.append({**a, "codigo_kit": codigo_kit, "__row": f"{i}.{j}"})
    return payload


# ── Upsert / apply ───────────────────────────────────────────────────────────


def apply_import(payload: ImportPayload) -> ImportSummary:
    summary = ImportSummary()
    with transaction.atomic():
        division_map = _upsert_divisions(payload.divisions, summary)
        kit_map = _upsert_kits(payload.kits, summary)
        _upsert_activities(payload.activities, kit_map, division_map, summary)
    return summary


def _upsert_divisions(rows, summary) -> dict[str, MasterFormat]:
    result: dict[str, MasterFormat] = {}
    for row in rows:
        code = _clean_str(row.get("division_code"))
        name = _clean_str(row.get("division_name"))
        if not code:
            summary.errors.append({
                "sheet": SHEET_DIVISIONS, "row": row.get("__row"),
                "message": "Falta division_code.",
            })
            continue
        obj, created = MasterFormat.objects.update_or_create(
            division_code=code,
            defaults={"division_name": name or code},
        )
        result[code] = obj
        if created:
            summary.divisions_created += 1
        else:
            summary.divisions_updated += 1
    return result


def _upsert_kits(rows, summary) -> dict[str, ActivityKit]:
    """Devuelve un diccionario indexado por codigo_kit → ActivityKit maestro."""
    result: dict[str, ActivityKit] = {}
    for row in rows:
        codigo = _clean_str(row.get("codigo_kit"))
        nombre = _clean_str(row.get("nombre"))
        descripcion = _clean_str(row.get("descripcion"))
        color = _clean_str(row.get("color")) or "#3b82f6"

        if not nombre and not codigo:
            summary.errors.append({
                "sheet": SHEET_KITS, "row": row.get("__row"),
                "message": "Un kit debe tener al menos codigo_kit o nombre.",
            })
            continue

        defaults = {
            "nombre": nombre or codigo,
            "descripcion": descripcion,
            "color": color,
        }
        if codigo:
            obj, created = ActivityKit.objects.update_or_create(
                codigo_kit=codigo, defaults=defaults,
            )
        else:
            existing = ActivityKit.objects.filter(
                nombre=nombre, proyecto__isnull=True,
            ).first()
            if existing:
                for k, v in defaults.items():
                    setattr(existing, k, v)
                existing.save()
                obj, created = existing, False
            else:
                obj = ActivityKit.objects.create(**defaults)
                created = True

        # Kit maestro: proyecto siempre NULL.
        if obj.proyecto_id is not None:
            obj.proyecto = None
            obj.save(update_fields=["proyecto"])

        result[codigo or obj.nombre] = obj
        if created:
            summary.kits_created += 1
        else:
            summary.kits_updated += 1
    return result


def _upsert_activities(rows, kit_map, division_map, summary):
    for row in rows:
        codigo_kit = _clean_str(row.get("codigo_kit"))
        codigo_act = _clean_str(row.get("codigo_actividad"))
        descripcion = _clean_str(row.get("descripcion"))
        unidad = _clean_str(row.get("unidad"))
        division_code = _clean_str(row.get("division_code"))

        row_ref = row.get("__row")

        if not codigo_act:
            summary.errors.append({
                "sheet": SHEET_ACTIVITIES, "row": row_ref,
                "message": "Falta codigo_actividad.",
            })
            continue
        if not codigo_kit:
            summary.errors.append({
                "sheet": SHEET_ACTIVITIES, "row": row_ref,
                "message": f"Actividad «{codigo_act}» sin codigo_kit.",
            })
            continue

        kit = kit_map.get(codigo_kit)
        if kit is None:
            kit = ActivityKit.objects.filter(
                codigo_kit=codigo_kit, proyecto__isnull=True,
            ).first()
            if kit is not None:
                kit_map[codigo_kit] = kit
        if kit is None:
            summary.errors.append({
                "sheet": SHEET_ACTIVITIES, "row": row_ref,
                "message": f"Kit «{codigo_kit}» no existe. Inclúyelo en la hoja Kits.",
            })
            continue

        division = division_map.get(division_code) if division_code else None
        if division is None and division_code:
            division = MasterFormat.objects.filter(division_code=division_code).first()
            if division:
                division_map[division_code] = division
        if division is None:
            summary.errors.append({
                "sheet": SHEET_ACTIVITIES, "row": row_ref,
                "message": (
                    f"División «{division_code or '(vacía)'}» no encontrada para "
                    f"actividad «{codigo_act}». Inclúyela en la hoja Divisiones."
                ),
            })
            continue

        try:
            cu_total = _to_decimal(row.get("cu_total"))
            material = _to_decimal(row.get("material"), default=Decimal("0"))
            mano_obra = _to_decimal(row.get("mano_obra"), default=Decimal("0"))
            equipo = _to_decimal(row.get("equipo"), default=Decimal("0"))
        except InvalidOperation as exc:
            summary.errors.append({
                "sheet": SHEET_ACTIVITIES, "row": row_ref,
                "message": f"Valor numérico inválido en actividad «{codigo_act}»: {exc}",
            })
            continue

        defaults = {
            "descripcion": descripcion,
            "unidad": unidad,
            "cu_total": cu_total,
            "material": material,
            "mano_obra": mano_obra,
            "equipo": equipo,
            "division": division,
        }

        existing = Activity.objects.filter(
            activity_kit=kit, codigo_actividad=codigo_act,
        ).first()
        if existing:
            for k, v in defaults.items():
                setattr(existing, k, v)
            existing.proyecto = None
            existing.save()
            summary.activities_updated += 1
        else:
            Activity.objects.create(
                codigo_actividad=codigo_act,
                activity_kit=kit,
                proyecto=None,
                **defaults,
            )
            summary.activities_created += 1


# ── Helpers ──────────────────────────────────────────────────────────────────


def _clean_str(value) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _to_decimal(value, default: Decimal | None = None) -> Decimal:
    if value is None or (isinstance(value, str) and not value.strip()):
        if default is not None:
            return default
        raise InvalidOperation("cu_total es obligatorio")
    if isinstance(value, Decimal):
        return value
    return Decimal(str(value).strip().replace(",", "."))


# ── Plantilla descargable ────────────────────────────────────────────────────


# ── Exportación de la biblioteca ─────────────────────────────────────────────


def _collect_library_snapshot() -> dict[str, list[dict[str, Any]]]:
    """Extrae los kits maestros y sus dependencias en filas listas para exportar."""
    kits = list(
        ActivityKit.objects.filter(proyecto__isnull=True)
        .prefetch_related("kit_activities__division")
        .order_by("codigo_kit", "nombre")
    )

    # Divisiones referenciadas por al menos una actividad de la biblioteca.
    used_division_ids = {
        act.division_id
        for kit in kits
        for act in kit.kit_activities.all()
        if act.division_id is not None
    }
    divisions = list(
        MasterFormat.objects.filter(id__in=used_division_ids).order_by("division_code")
    )

    div_rows = [{"division_code": d.division_code, "division_name": d.division_name} for d in divisions]
    kit_rows = [
        {
            "codigo_kit": k.codigo_kit or "",
            "nombre": k.nombre,
            "descripcion": k.descripcion or "",
            "color": k.color or "",
        }
        for k in kits
    ]
    activity_rows = []
    for kit in kits:
        for act in kit.kit_activities.all().order_by("codigo_actividad"):
            activity_rows.append({
                "codigo_kit": kit.codigo_kit or kit.nombre,
                "codigo_actividad": act.codigo_actividad,
                "descripcion": act.descripcion,
                "unidad": act.unidad,
                "cu_total": str(act.cu_total),
                "material": str(act.material),
                "mano_obra": str(act.mano_obra),
                "equipo": str(act.equipo),
                "division_code": act.division.division_code if act.division_id else "",
            })

    return {"divisions": div_rows, "kits": kit_rows, "activities": activity_rows}


def export_library_xlsx() -> bytes:
    snapshot = _collect_library_snapshot()
    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_DIVISIONS
    ws.append(DIVISION_HEADERS)
    for row in snapshot["divisions"]:
        ws.append([row[h] for h in DIVISION_HEADERS])

    ws2 = wb.create_sheet(SHEET_KITS)
    ws2.append(KIT_HEADERS)
    for row in snapshot["kits"]:
        ws2.append([row[h] for h in KIT_HEADERS])

    ws3 = wb.create_sheet(SHEET_ACTIVITIES)
    ws3.append(ACTIVITY_HEADERS)
    for row in snapshot["activities"]:
        ws3.append([row[h] for h in ACTIVITY_HEADERS])

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def export_library_json() -> bytes:
    snapshot = _collect_library_snapshot()
    # Formato anidado para round-trip con _parse_json.
    kits_by_code: dict[str, dict[str, Any]] = {}
    for k in snapshot["kits"]:
        entry = dict(k)
        entry["actividades"] = []
        kits_by_code[k["codigo_kit"] or k["nombre"]] = entry
    for a in snapshot["activities"]:
        key = a["codigo_kit"]
        if key in kits_by_code:
            act = {k: v for k, v in a.items() if k != "codigo_kit"}
            kits_by_code[key]["actividades"].append(act)

    payload = {
        "divisiones": snapshot["divisions"],
        "kits": list(kits_by_code.values()),
    }
    return json.dumps(payload, ensure_ascii=False, indent=2).encode("utf-8")


def export_library_csv_zip() -> bytes:
    """Devuelve un ZIP con tres CSVs (divisiones, kits, actividades)."""
    snapshot = _collect_library_snapshot()
    buf = BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
        for name, headers, rows in [
            ("divisiones.csv", DIVISION_HEADERS, snapshot["divisions"]),
            ("kits.csv", KIT_HEADERS, snapshot["kits"]),
            ("actividades.csv", ACTIVITY_HEADERS, snapshot["activities"]),
        ]:
            sio = StringIO()
            writer = csv.DictWriter(sio, fieldnames=headers)
            writer.writeheader()
            for row in rows:
                writer.writerow({h: row.get(h, "") for h in headers})
            zf.writestr(name, sio.getvalue().encode("utf-8-sig"))
    return buf.getvalue()


def build_template_workbook() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_DIVISIONS
    ws.append(DIVISION_HEADERS)
    ws.append(["03", "Concrete"])
    ws.append(["04", "Masonry"])

    ws2 = wb.create_sheet(SHEET_KITS)
    ws2.append(KIT_HEADERS)
    ws2.append(["KIT-001", "Fundaciones", "Kit ejemplo de fundaciones", "#3b82f6"])
    ws2.append(["KIT-002", "Muros", "Kit ejemplo de muros", "#1a7a52"])

    ws3 = wb.create_sheet(SHEET_ACTIVITIES)
    ws3.append(ACTIVITY_HEADERS)
    ws3.append(["KIT-001", "A-001", "Excavación manual", "m3", 50, 15, 25, 10, "03"])
    ws3.append(["KIT-001", "A-002", "Concreto 3000 psi", "m3", 220, 140, 60, 20, "03"])
    ws3.append(["KIT-002", "A-003", "Muro bloque 8\"", "m2", 45, 25, 15, 5, "04"])

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
