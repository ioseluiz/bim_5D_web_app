"""Tests para costs.importers: parseo y upsert de la biblioteca de kits."""
from __future__ import annotations

import json
from decimal import Decimal
from io import BytesIO

import pytest
from django.core.files.uploadedfile import SimpleUploadedFile
from openpyxl import Workbook

from costs.importers import (
    ACTIVITY_HEADERS,
    DIVISION_HEADERS,
    KIT_HEADERS,
    SHEET_ACTIVITIES,
    SHEET_DIVISIONS,
    SHEET_KITS,
    ImportError_,
    apply_import,
    build_template_workbook,
    parse_file,
)
from costs.models import Activity, ActivityKit, MasterFormat


pytestmark = pytest.mark.django_db


def _make_xlsx(divisions=None, kits=None, activities=None) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_DIVISIONS
    ws.append(DIVISION_HEADERS)
    for row in divisions or []:
        ws.append(row)

    ws2 = wb.create_sheet(SHEET_KITS)
    ws2.append(KIT_HEADERS)
    for row in kits or []:
        ws2.append(row)

    ws3 = wb.create_sheet(SHEET_ACTIVITIES)
    ws3.append(ACTIVITY_HEADERS)
    for row in activities or []:
        ws3.append(row)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _upload(content: bytes, name: str) -> SimpleUploadedFile:
    return SimpleUploadedFile(name, content)


# ── Parseo ───────────────────────────────────────────────────────────────────


def test_parse_xlsx_reads_three_sheets():
    content = _make_xlsx(
        divisions=[["03", "Concrete"]],
        kits=[["KIT-A", "Kit A", "desc", "#111111"]],
        activities=[["KIT-A", "ACT-1", "Excavación", "m3", 50, 15, 25, 10, "03"]],
    )
    payload = parse_file(_upload(content, "biblioteca.xlsx"))
    assert len(payload.divisions) == 1
    assert len(payload.kits) == 1
    assert len(payload.activities) == 1
    assert payload.divisions[0]["division_code"] == "03"


def test_parse_csv_detects_activities_by_headers():
    csv = (
        "codigo_kit,codigo_actividad,descripcion,unidad,cu_total,material,mano_obra,equipo,division_code\n"
        "KIT-A,ACT-1,Excavación,m3,50,15,25,10,03\n"
    )
    payload = parse_file(_upload(csv.encode("utf-8"), "actividades.csv"))
    assert payload.activities and not payload.divisions and not payload.kits


def test_parse_json_flattens_nested_activities():
    data = {
        "divisiones": [{"division_code": "03", "division_name": "Concrete"}],
        "kits": [
            {
                "codigo_kit": "KIT-A",
                "nombre": "Kit A",
                "actividades": [
                    {"codigo_actividad": "ACT-1", "descripcion": "Excavación",
                     "unidad": "m3", "cu_total": 50, "division_code": "03"},
                ],
            }
        ],
    }
    payload = parse_file(_upload(json.dumps(data).encode("utf-8"), "biblioteca.json"))
    assert len(payload.kits) == 1
    assert len(payload.activities) == 1
    assert payload.activities[0]["codigo_kit"] == "KIT-A"


def test_parse_unsupported_extension_raises():
    with pytest.raises(ImportError_):
        parse_file(_upload(b"foo", "biblioteca.txt"))


def test_template_workbook_is_valid_xlsx():
    content = build_template_workbook()
    payload = parse_file(_upload(content, "plantilla.xlsx"))
    assert payload.divisions and payload.kits and payload.activities


# ── Upsert / apply_import ────────────────────────────────────────────────────


def test_apply_import_creates_full_library():
    content = _make_xlsx(
        divisions=[["03", "Concrete"], ["04", "Masonry"]],
        kits=[
            ["KIT-A", "Kit A", "desc A", "#111111"],
            ["KIT-B", "Kit B", "desc B", "#222222"],
        ],
        activities=[
            ["KIT-A", "ACT-1", "Excavación", "m3", 50, 15, 25, 10, "03"],
            ["KIT-A", "ACT-2", "Concreto",   "m3", 220, 140, 60, 20, "03"],
            ["KIT-B", "ACT-3", "Muro",       "m2", 45, 25, 15, 5, "04"],
        ],
    )
    summary = apply_import(parse_file(_upload(content, "lib.xlsx")))

    assert summary.divisions_created == 2
    assert summary.kits_created == 2
    assert summary.activities_created == 3
    assert summary.errors == []

    assert MasterFormat.objects.count() == 2
    assert ActivityKit.objects.filter(proyecto__isnull=True).count() == 2
    assert Activity.objects.count() == 3
    kit_a = ActivityKit.objects.get(codigo_kit="KIT-A")
    assert kit_a.kit_activities.count() == 2
    assert kit_a.kit_activities.get(codigo_actividad="ACT-2").cu_total == Decimal("220")


def test_apply_import_is_idempotent_and_updates_on_reimport():
    content = _make_xlsx(
        divisions=[["03", "Concrete"]],
        kits=[["KIT-A", "Kit A", "desc", "#111111"]],
        activities=[["KIT-A", "ACT-1", "Excavación", "m3", 50, 15, 25, 10, "03"]],
    )
    apply_import(parse_file(_upload(content, "lib.xlsx")))

    updated = _make_xlsx(
        divisions=[["03", "Concrete"]],
        kits=[["KIT-A", "Kit A actualizado", "nueva desc", "#111111"]],
        activities=[["KIT-A", "ACT-1", "Excavación mejor", "m3", 99, 20, 60, 19, "03"]],
    )
    summary = apply_import(parse_file(_upload(updated, "lib.xlsx")))

    assert summary.kits_created == 0
    assert summary.kits_updated == 1
    assert summary.activities_created == 0
    assert summary.activities_updated == 1

    kit = ActivityKit.objects.get(codigo_kit="KIT-A")
    assert kit.nombre == "Kit A actualizado"
    act = kit.kit_activities.get(codigo_actividad="ACT-1")
    assert act.cu_total == Decimal("99")
    assert act.descripcion == "Excavación mejor"


def test_missing_division_reports_error_and_skips_activity():
    content = _make_xlsx(
        kits=[["KIT-A", "Kit A", "desc", "#111111"]],
        activities=[["KIT-A", "ACT-1", "Sin división", "m3", 50, 0, 0, 0, "99"]],
    )
    summary = apply_import(parse_file(_upload(content, "lib.xlsx")))
    assert summary.activities_created == 0
    assert any("División" in e["message"] for e in summary.errors)


def test_missing_kit_reference_reports_error():
    content = _make_xlsx(
        divisions=[["03", "Concrete"]],
        activities=[["KIT-FANTASMA", "ACT-1", "desc", "m3", 10, 0, 0, 0, "03"]],
    )
    summary = apply_import(parse_file(_upload(content, "lib.xlsx")))
    assert summary.activities_created == 0
    assert any("Kit" in e["message"] for e in summary.errors)


def test_kit_without_codigo_uses_nombre_fallback():
    content = _make_xlsx(
        divisions=[["03", "Concrete"]],
        kits=[[None, "Kit Sin Código", "desc", "#111111"]],
    )
    summary = apply_import(parse_file(_upload(content, "lib.xlsx")))
    assert summary.kits_created == 1
    assert ActivityKit.objects.filter(nombre="Kit Sin Código").exists()
