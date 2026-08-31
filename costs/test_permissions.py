"""Tests de permisos: staff vs. usuario normal sobre la biblioteca compartida."""
from __future__ import annotations

from decimal import Decimal
from io import BytesIO

import pytest
from django.contrib.auth import get_user_model
from openpyxl import Workbook
from rest_framework.test import APIClient

from bim.models import Project
from costs.importers import (
    ACTIVITY_HEADERS,
    DIVISION_HEADERS,
    KIT_HEADERS,
    SHEET_ACTIVITIES,
    SHEET_DIVISIONS,
    SHEET_KITS,
)
from costs.models import Activity, ActivityKit, MasterFormat


pytestmark = pytest.mark.django_db
User = get_user_model()


# ── Fixtures ─────────────────────────────────────────────────────────────────


@pytest.fixture
def staff_user(db):
    return User.objects.create_user(
        username="admin", email="admin@example.com", password="x", is_staff=True,
    )


@pytest.fixture
def regular_user(db):
    return User.objects.create_user(
        username="ana", email="ana@example.com", password="x",
    )


@pytest.fixture
def other_user(db):
    return User.objects.create_user(
        username="pedro", email="pedro@example.com", password="x",
    )


@pytest.fixture
def staff_client(staff_user):
    c = APIClient()
    c.force_authenticate(staff_user)
    return c


@pytest.fixture
def user_client(regular_user):
    c = APIClient()
    c.force_authenticate(regular_user)
    return c


@pytest.fixture
def anon_client():
    return APIClient()


@pytest.fixture
def division(db):
    return MasterFormat.objects.create(division_code="03", division_name="Concrete")


@pytest.fixture
def master_kit(division):
    kit = ActivityKit.objects.create(codigo_kit="KIT-M", nombre="Kit maestro", proyecto=None)
    Activity.objects.create(
        codigo_actividad="A-1", descripcion="Excavación", unidad="m3",
        cu_total=Decimal("50"), division=division, activity_kit=kit,
    )
    return kit


@pytest.fixture
def project(regular_user):
    return Project.objects.create(owner=regular_user, nombre="Proyecto Ana")


def _minimal_xlsx() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_DIVISIONS
    ws.append(DIVISION_HEADERS)
    ws.append(["03", "Concrete"])
    ws2 = wb.create_sheet(SHEET_KITS)
    ws2.append(KIT_HEADERS)
    ws2.append(["KIT-X", "Kit X", "d", "#111111"])
    ws3 = wb.create_sheet(SHEET_ACTIVITIES)
    ws3.append(ACTIVITY_HEADERS)
    ws3.append(["KIT-X", "A-1", "Excavación", "m3", 50, 15, 25, 10, "03"])
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── Lectura: todos los autenticados pueden ver la biblioteca ─────────────────


def test_regular_user_can_list_master_kits(user_client, master_kit):
    resp = user_client.get("/api/activity-kits/")
    assert resp.status_code == 200
    assert any(k["codigo_kit"] == "KIT-M" for k in resp.json())


def test_anonymous_cannot_list_master_kits(anon_client, master_kit):
    resp = anon_client.get("/api/activity-kits/")
    assert resp.status_code in (401, 403)


# ── Escritura: sólo staff puede modificar la biblioteca ──────────────────────


def test_regular_user_cannot_create_master_kit(user_client):
    resp = user_client.post("/api/activity-kits/", {"nombre": "Hack", "codigo_kit": "HACK"})
    assert resp.status_code == 403
    assert not ActivityKit.objects.filter(codigo_kit="HACK").exists()


def test_regular_user_cannot_update_master_kit(user_client, master_kit):
    resp = user_client.patch(f"/api/activity-kits/{master_kit.id}/", {"nombre": "Hackeado"})
    assert resp.status_code == 403
    master_kit.refresh_from_db()
    assert master_kit.nombre == "Kit maestro"


def test_regular_user_cannot_delete_master_kit(user_client, master_kit):
    resp = user_client.delete(f"/api/activity-kits/{master_kit.id}/")
    assert resp.status_code == 403
    assert ActivityKit.objects.filter(id=master_kit.id).exists()


def test_staff_can_update_master_kit(staff_client, master_kit):
    resp = staff_client.patch(f"/api/activity-kits/{master_kit.id}/", {"nombre": "Renombrado"})
    assert resp.status_code == 200
    master_kit.refresh_from_db()
    assert master_kit.nombre == "Renombrado"


# ── import_library: sólo staff ───────────────────────────────────────────────


def test_regular_user_cannot_import_library(user_client):
    content = _minimal_xlsx()
    resp = user_client.post(
        "/api/activity-kits/import_library/",
        {"files": BytesIO(content)},
        format="multipart",
    )
    # Los archivos hay que enviarlos como SimpleUploadedFile; usamos APIClient con
    # multipart. Como no adjuntamos archivo válido, verificamos rechazo por permiso
    # primero (403) sobre lo que sea.
    assert resp.status_code == 403
    assert not ActivityKit.objects.filter(codigo_kit="KIT-X").exists()


def test_staff_can_import_library(staff_client):
    from django.core.files.uploadedfile import SimpleUploadedFile
    upload = SimpleUploadedFile(
        "biblioteca.xlsx", _minimal_xlsx(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    resp = staff_client.post(
        "/api/activity-kits/import_library/", {"files": upload}, format="multipart",
    )
    assert resp.status_code == 200, resp.content
    assert ActivityKit.objects.filter(codigo_kit="KIT-X").exists()


def test_anonymous_cannot_hit_import_library(anon_client):
    resp = anon_client.post("/api/activity-kits/import_library/")
    assert resp.status_code in (401, 403)


# ── copy_to_project: cualquier autenticado con su propio proyecto ────────────


def test_regular_user_can_copy_master_to_own_project(user_client, master_kit, project):
    resp = user_client.post(
        f"/api/activity-kits/{master_kit.id}/copy_to_project/",
        {"proyecto": project.id},
    )
    assert resp.status_code == 201
    # La copia hereda nombre y actividades, pero codigo_kit queda NULL (es unique global).
    copied = ActivityKit.objects.filter(proyecto=project, nombre=master_kit.nombre).first()
    assert copied is not None
    assert copied.codigo_kit is None
    assert copied.kit_activities.count() == master_kit.kit_activities.count()


def test_regular_user_cannot_copy_to_another_users_project(user_client, master_kit, other_user):
    other_project = Project.objects.create(owner=other_user, nombre="Otro")
    resp = user_client.post(
        f"/api/activity-kits/{master_kit.id}/copy_to_project/",
        {"proyecto": other_project.id},
    )
    assert resp.status_code == 403
    assert not ActivityKit.objects.filter(proyecto=other_project).exists()


# ── Kits de proyecto: sólo el owner ──────────────────────────────────────────


def test_owner_can_edit_own_project_kit(user_client, project, division):
    kit = ActivityKit.objects.create(nombre="Kit ana", proyecto=project)
    resp = user_client.patch(f"/api/activity-kits/{kit.id}/", {"nombre": "Renombrado por ana"})
    assert resp.status_code == 200


def test_other_user_cannot_see_or_edit_someone_elses_project_kit(other_user, project):
    kit = ActivityKit.objects.create(nombre="Kit ana", proyecto=project)
    c = APIClient()
    c.force_authenticate(other_user)
    # No debe siquiera aparecer en el queryset del otro usuario.
    resp = c.patch(f"/api/activity-kits/{kit.id}/", {"nombre": "Hack"})
    assert resp.status_code == 404


# ── MasterFormat: staff-write, read-all ──────────────────────────────────────


def test_regular_user_cannot_create_masterformat(user_client):
    resp = user_client.post("/api/masterformat/", {"division_code": "99", "division_name": "X"})
    assert resp.status_code == 403


def test_staff_can_create_masterformat(staff_client):
    resp = staff_client.post("/api/masterformat/", {"division_code": "99", "division_name": "X"})
    assert resp.status_code == 201


def test_anonymous_cannot_read_masterformat(anon_client, division):
    resp = anon_client.get("/api/masterformat/")
    assert resp.status_code in (401, 403)


def test_regular_user_can_read_masterformat(user_client, division):
    resp = user_client.get("/api/masterformat/")
    assert resp.status_code == 200


# ── template descarga: sólo autenticados ─────────────────────────────────────


def test_regular_user_can_download_template(user_client):
    resp = user_client.get("/api/activity-kits/template/")
    assert resp.status_code == 200
    assert resp["Content-Type"].startswith("application/vnd.openxmlformats")


def test_anonymous_cannot_download_template(anon_client):
    resp = anon_client.get("/api/activity-kits/template/")
    assert resp.status_code in (401, 403)
